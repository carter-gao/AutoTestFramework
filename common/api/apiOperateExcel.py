#!/usr/bin/env python
# coding:utf-8

# @Author:  Carter.Gao
# @Email:   gaoxx9527@163.com
# @Date:    2019/8/21 15:45
# @IDE:     PyCharm
# @About:   封装EXCEL操作

import os
from time import strftime
from shutil import move
import openpyxl
from openpyxl.chart.series import DataPoint
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from common import constant
from common.logger import Logger


# ----------------------------------------------------------------------------------------
# 封装测试结果处理类：
# 1.在每次执行全量测试之前，先调用BackupOrNewResultFile类的备份方法备份
# 2.再调用BackupOrNewResultFile类的新建文件方法，创建本次测试的回写结果文件
# 3.回写类在测试用例基类的setUpClass方法内实例化，这样每个接口类只有一次加载excel操作，可降低内存占用
# 4.回写方法在setUp方法、tearDown方法，以及测试用例内部调用，可灵活使用
# 5.预期结果及判断结果方法在测试用例基类内部封装好的断言方法中调用，断言方法在用例内部调用
# 6.在测试用例基类的tearDownClass方法内调用save_excel()方法，一次性保存当前接口所有已回写的测试结果
# 7.分析类AnalyzeResults在测试执行完毕，生成回写文件后实例化，调用exec_analysis()处理测试结果
# ----------------------------------------------------------------------------------------


# 给定EXCEL回写常量
API_NAME_COL = 1            # 接口名称所在列
API_URL_COL = 2             # 接口地址所在列
CASE_NUMBER_COL = 3         # 用例编号所在列
CASE_NAME_COL = 4           # 用例名称所在列
JUDGEMENT_RESULT_COL = 5    # 判定结果所在列
EXPECTED_RESULT_COL = 6     # 预期结果所在列
COMPARE_RESULT_COL = 7      # 比对结果所在列
RESPONSE_COL = 8            # 返回结果所在列
CASE_DATA_COL = 9           # 测试数据所在列


class BackupOrNewFile:
    """
    封装或新建测试结果excel
    一般在本次测试测试启动前，先调用backup_result_file备份上一次的文件
    在调用create_result_file创建本次文件
    """
    def __init__(self):
        # 引用日志类
        self._log = Logger('备份或新建API测试结果文件').get_logger()

        # 获取excel绝对路径
        self._path = constant.api_result_excel_path

    def backup_result_file(self):
        """
        用于每次执行完整测试前备份excel文件
        需要在执行测试前单独调用此函数先备份，再调用create_result_file函数创建新的excel文件
        :return: None
        """
        if os.path.exists(self._path):
            current_time = strftime('%Y-%m-%d %H-%M-%S')
            backup_file_name = '{} backup at {}.xlsx'.format(self._path.split('.')[0], current_time)
            move(self._path, backup_file_name)
            self._log.info('成功备份测试结果文件：{}'.format(backup_file_name))
        else:
            self._log.info('不存在需要备份的文件。')

    def create_result_file(self):
        """
        用于每次备份完成后调用，创建新的excel文件
        :return: None
        """
        if not os.path.exists(self._path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'results'
            self._set_row_one_values(ws)
            self._set_row_one_width_and_height(ws)
            self._set_row_one_styles(ws)
            # 冻结首行
            ws.freeze_panes = 'A2'
            wb.save(self._path)
            self._log.info('成功创建测试结果文件：{}'.format(self._path))
        else:
            self._log.info('{}文件已存在，请先做好备份。'.format(os.path.basename(self._path)))

    @staticmethod
    def _set_row_one_values(ws):
        """
        设置excel标题栏的值
        :param ws: worksheet对象
        :return: None
        """
        ws.cell(1, API_NAME_COL).value = '接口名称'
        ws.cell(1, API_URL_COL).value = '接口URL'
        ws.cell(1, CASE_NUMBER_COL).value = '用例编号'
        ws.cell(1, CASE_NAME_COL).value = '用例名称'
        ws.cell(1, JUDGEMENT_RESULT_COL).value = '判定结果'
        ws.cell(1, EXPECTED_RESULT_COL).value = '预期结果'
        ws.cell(1, COMPARE_RESULT_COL).value = '比对结果'
        ws.cell(1, RESPONSE_COL).value = '响应结果'
        ws.cell(1, CASE_DATA_COL).value = '测试数据'

    @staticmethod
    def _set_row_one_width_and_height(ws):
        """
        设置excel标题栏列宽和行高
        :param ws: worksheet对象
        :return: None
        """
        ws.column_dimensions[get_column_letter(API_NAME_COL)].width = 30.0
        ws.column_dimensions[get_column_letter(API_URL_COL)].width = 30.0
        ws.column_dimensions[get_column_letter(CASE_NUMBER_COL)].width = 12.0
        ws.column_dimensions[get_column_letter(CASE_NAME_COL)].width = 30.0
        ws.column_dimensions[get_column_letter(JUDGEMENT_RESULT_COL)].width = 12.0
        ws.column_dimensions[get_column_letter(EXPECTED_RESULT_COL)].width = 40.0
        ws.column_dimensions[get_column_letter(COMPARE_RESULT_COL)].width = 40.0
        ws.column_dimensions[get_column_letter(RESPONSE_COL)].width = 60.0
        ws.column_dimensions[get_column_letter(CASE_DATA_COL)].width = 40.0
        ws.row_dimensions[1].height = 25.0

    @staticmethod
    def _set_row_one_styles(ws):
        """
        设置excel标题栏样式
        :param ws: worksheet对象
        :return: None
        """
        # 定义字体样式，填充颜色，边框
        font = Font(name=u'宋体', size=14, bold=True)
        fill = PatternFill(fill_type='solid', fgColor="32CD32")
        border = Border(
            left=Side(border_style='double', color='FF000000'),
            right=Side(border_style='double', color='FF000000'),
            top=Side(border_style='double', color='FF000000'),
            bottom=Side(border_style='double', color='FF000000')
        )
        for col in range(CASE_DATA_COL):
            ws.cell(1, col + 1).font = font
            ws.cell(1, col + 1).fill = fill
            ws.cell(1, col + 1).border = border


class BackFillToExcel:
    """
    测试结果回写
    """
    def __init__(self):
        # 引用日志类
        self._log = Logger('API测试结果回写').get_logger()

        # 获取excel绝对路径
        self._path = constant.api_result_excel_path

        # 初始化判断，若excel文件不存在，就创建
        if not os.path.exists(self._path):
            BackupOrNewFile().create_result_file()

        # 加载excel，并转到指定的sheet页，返回WorkSheet对象
        self._wb = openpyxl.load_workbook(self._path)
        self._ws = self._wb[self._wb.sheetnames[0]]

        self._log.debug('成功加载测试结果文件：{}'.format(os.path.basename(self._path)))
        self._log.debug('成功定位到回写数据表：{}'.format(self._wb.sheetnames[0]))

    def save_excel(self):
        """
        保存当前用例所有回填结果
        :return: None
        """
        try:
            # self._merge_cells_before_save()
            self._wb.save(self._path)
        except PermissionError as e:
            self._log.error('保存失败！！EXCEL文件被打开，请关闭后重新执行测试：{}'.format(e))
        else:
            self._log.debug('成功保存当前用例测试结果。')

    def fill_api_name(self, api_name):
        """
        回写接口名称
        :param api_name: 接口名
        :return: None
        """
        self._ws.cell(self._ws.max_row, API_NAME_COL).value = api_name
        self._log.debug('成功回写当前接口名称：{}'.format(api_name))

    def fill_api_url(self, api_url):
        """
        回写接口URL
        :param api_url: URL
        :return: None
        """
        self._ws.cell(self._ws.max_row, API_URL_COL).value = api_url
        self._log.debug('成功回写当前接口URL：{}'.format(api_url))

    def fill_case_number(self, case_number):
        """
        回写用例编号
        注：此函数在基类的setUp方法内第一个调用，先确定回写行数，其他回写函数就可以最大行数为准写入
        :param case_number: 用例编号
        :return: None
        """
        # 第一个调用，所以行数是 max_row + 1
        self._ws.cell(self._ws.max_row + 1, CASE_NUMBER_COL).value = case_number
        self._log.debug('成功回写当前用例编号为{}，并确定当前用例回写行是第{}行'.format(case_number, self._ws.max_row))

    def fill_case_name(self, case_name):
        """
        回写用例名称
        :param case_name: 用例名
        :return: None
        """
        self._ws.cell(self._ws.max_row, CASE_NAME_COL).value = case_name
        self._log.debug('成功回写当前用例名称：{}'.format(case_name))

    def fill_judgement_result(self, result=1):
        """
        回写判定结果
        :param result: 0 失败，1 成功
        :return: None
        """
        if result:
            self._ws.cell(self._ws.max_row, JUDGEMENT_RESULT_COL).value = 'SUCCESS'
            self._log.debug('默认回写当前用例执行判定结果为：SUCCESS')
        else:
            self._ws.cell(self._ws.max_row, JUDGEMENT_RESULT_COL).value = 'FAILURE'
            self._log.debug('当前用例断言失败或执行失败，改写判定结果为：FAILURE')
            self._set_color_if_failure(self._ws.max_row)

    def fill_excepted(self, excepted_result):
        """
        回写预期结果
        :param excepted_result: 预期结果
        :return: None
        """
        self._ws.cell(self._ws.max_row, EXPECTED_RESULT_COL).value = str(excepted_result)
        self._log.debug('成功回写当前用例预期结果：{}'.format(excepted_result))

    def fill_compare_result(self, compare_result):
        """
        回写比对结果
        :param compare_result: 比对结果列表
        :return: None
        """
        if not compare_result:
            self._ws.cell(self._ws.max_row, COMPARE_RESULT_COL).value = '比对结果一致'
        else:
            self._ws.cell(self._ws.max_row, COMPARE_RESULT_COL).value = str(compare_result)
            self._log.debug('成功回写当前用例比对结果：{}'.format(compare_result))

    def fill_response(self, response):
        """
        回写返回JSON
        :param response: JSON
        :return: None
        """
        self._ws.cell(self._ws.max_row, RESPONSE_COL).value = str(response)
        self._log.debug('成功回写当前用例返回结果：{}'.format(response))

    def fill_test_data(self, curr_case_data):
        """
        回写本条用例测试数据
        :param curr_case_data: 测试数据
        :return: None
        """
        self._ws.cell(self._ws.max_row, CASE_DATA_COL).value = str(curr_case_data)
        self._log.debug('成功回写当前用例测试数据：{}'.format(curr_case_data))

    def _set_color_if_failure(self, row):
        """
        把执行失败的用例背景色填充为红色
        :param row: 行数
        :return: None
        """
        fill = PatternFill(fill_type='solid', fgColor="FF0000")
        for col in range(API_NAME_COL, CASE_DATA_COL + 1):
            self._ws.cell(row, col).fill = fill
        self._log.debug('当前用例执行失败标记为红色！')

    def _merge_cells_before_save(self):
        """
        在保存之前调用，用于合并单元格（用例名和URL）
        :return: None
        """
        # 确定最小合并行
        min_row_ = self._ws.max_row - self._ws.cell(self._ws.max_row, CASE_NUMBER_COL).value + 1
        # 合并单元格
        self._ws.merge_cells(start_row=min_row_, start_column=API_NAME_COL,
                             end_row=self._ws.max_row, end_column=API_NAME_COL)
        self._ws.merge_cells(start_row=min_row_, start_column=API_URL_COL,
                             end_row=self._ws.max_row, end_column=API_URL_COL)
        # 设置垂直居中
        align = Alignment(vertical='center')
        self._ws.cell(min_row_, API_NAME_COL).alignment = align
        self._ws.cell(min_row_, API_URL_COL).alignment = align


class AnalyzeResults:
    """
    分析处理测试结果
    """
    def __init__(self):
        # 引用日志类
        self._log = Logger('分析处理API测试结果').get_logger()

        # 获取excel路径
        self._path = constant.api_result_excel_path

        # 加载excel，返回WorkBook对象
        self._wb = openpyxl.load_workbook(self._path)
        self._log.info('成功加载测试结果文件：{}'.format(os.path.basename(self._path)))

        # 创建分析结果sheet页
        self._wb.create_sheet('analysis', index=0)
        self._log.info('成功创建分析结果sheet页：analysis')

        # 定义一些变量
        self._case_num = 0      # 用例总数
        self._case_failure = 0  # 失败用例数
        self._api_num = 0       # 接口总数
        self._api_failure = 0   # 失败接口数

    def exec_analysis(self):
        """
        执行分析计算
        :return: None
        """
        self._log.info('开始分析测试结果：')
        self._get_api_num()
        self._get_api_failure_num()
        self._get_case_num()
        self._get_case_failure_num()
        self._set_known_cells_value()
        result_api = self._statistical_failure_interfaces()
        first_max_row = len(result_api) + 4
        self._fill_failure_interfaces(result_api)
        self._set_unknown_cells_value(first_max_row)
        second_max_row = first_max_row + 3
        result_case = self._statistical_failure_cases()
        self._fill_failure_cases(result_case, second_max_row)
        self._draw_pie_charts()
        self._draw_bar_chart(first_max_row)
        self._log.info('测试结果分析完毕.')
        self._save_excel()

    def _save_excel(self):
        """
        保存分析结果
        :return: None
        """
        self._wb.save(self._path)
        self._log.info('成功保存，请打开excel文件analysis数据表查看！')

    def _statistical_failure_interfaces(self):
        """
        统计失败接口
        :return: None
        """
        ws = self._wb['results']
        self._log.info('正在提取失败接口相关数据...')
        # 先找到所有接口第一个和最后一个用例所在行，生成两个列表
        first_rows = []
        last_rows = []
        for i in range(2, ws.max_row + 1):
            if ws.cell(i, CASE_NUMBER_COL).value == 1:
                first_rows.append(i)
                if i != 2:
                    last_rows.append(i - 1)
        last_rows.append(ws.max_row)
        # 计算失败接口数据
        result_all = []
        for i, j in zip(first_rows, last_rows):
            result_single = {}
            failure_num = 0
            for row in range(i, j + 1):
                if ws.cell(row, JUDGEMENT_RESULT_COL).value == 'FAILURE':
                    failure_num += 1
            if failure_num:
                result_single['api_name'] = ws.cell(i, API_NAME_COL).value
                result_single['case_num'] = j - i + 1
                result_single['failure_num'] = failure_num
                result_single['success_num'] = j - i + 1 - failure_num
                result_all.append(result_single)
        self._log.info('提取成功：{}'.format(result_all))
        return result_all

    def _fill_failure_interfaces(self, data):
        """
        写入失败接口相关数据
        :param data: 传入数据
        :return: None
        """
        ws = self._wb['analysis']
        self._log.info('正在写入提取结果...')
        rows = [row for row in range(4, len(data) + 4)]
        for d, row in zip(data, rows):
            ws.cell(row, 1).value = d['api_name']
            ws.cell(row, 2).value = d['success_num']
            ws.cell(row, 3).value = d['failure_num']
            ws.cell(row, 4).value = d['case_num']
        self._log.info('写入完毕.')

    def _statistical_failure_cases(self):
        """
        统计失败用例
        :return: None
        """
        ws = self._wb['results']
        self._log.info('正在提取失败用例相关数据...')
        case_failures = []
        for i in range(2, ws.max_row + 1):
            case_failure = []
            if ws.cell(i, JUDGEMENT_RESULT_COL).value == 'FAILURE':
                case_failure.append(ws.cell(i, API_NAME_COL).value)
                case_failure.append(ws.cell(i, CASE_NAME_COL).value)
                case_failure.append(ws.cell(i, CASE_NUMBER_COL).value)
                # 计算当前行范围，用作超链接目标
                row_region = '{}{}:{}{}'.format(
                    get_column_letter(API_NAME_COL), i,
                    get_column_letter(CASE_DATA_COL), i
                )
                case_failure.append(row_region)
            if case_failure:
                case_failures.append(case_failure)
        self._log.info('提取成功：{}'.format(case_failures))
        return case_failures

    def _fill_failure_cases(self, data, fill_row):
        """
        写入失败用例相关数据
        :param data: 用例测试数据
        :param fill_row: 写入行
        :return: None
        """
        ws = self._wb['analysis']
        self._log.info('正在写入提取结果...')
        rows = [row for row in range(fill_row, fill_row + len(data))]
        for d, row in zip(data, rows):
            ws.cell(row, 1).value = d[0]
            ws.cell(row, 2).value = d[1]
            ws.cell(row, 3).value = d[2]
            ws.cell(row, 4).value = '点我查看失败原因'
            ws.cell(row, 4).hyperlink = ('{}#results!{}'.format(constant.API_TEST_RESULT_EXCEL, d[3]))
        self._log.info('写入完毕.')

    def _draw_pie_charts(self):
        """
        画两个饼图
        :return: None
        """
        ws = self._wb['analysis']
        # 设置单元格值，饼图引用
        ws['G3'] = '失败'
        ws['G4'] = '通过'
        ws['H3'] = self._api_failure
        ws['H4'] = self._api_num - self._api_failure
        ws['N3'] = '失败'
        ws['N4'] = '通过'
        ws['O3'] = self._case_failure
        ws['O4'] = self._case_num - self._case_failure

        # 画接口饼图
        pie = PieChart()
        labels = Reference(ws, min_col=7, min_row=3, max_row=4)
        data = Reference(ws, min_col=8, min_row=2, max_row=4)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "接口执行情况"
        slice_ = DataPoint(idx=0, explosion=10)
        pie.series[0].data_points = [slice_]
        ws.add_chart(pie, "F1")
        pie.height = 9.5
        pie.width = 13
        self._log.info('已生成接口执行情况饼图.')

        # 画用例饼图
        pie2 = PieChart()
        labels2 = Reference(ws, min_col=14, min_row=3, max_row=4)
        data2 = Reference(ws, min_col=15, min_row=2, max_row=4)
        pie2.add_data(data2, titles_from_data=True)
        pie2.set_categories(labels2)
        pie2.title = "用例执行情况"
        slice2_ = DataPoint(idx=0, explosion=10)
        pie2.series[0].data_points = [slice2_]
        ws.add_chart(pie2, "M1")
        pie2.height = 9.5
        pie2.width = 13
        self._log.info('已生成用例执行情况饼图.')

    def _draw_bar_chart(self, row_):
        """
        画垂直条形图
        :param row_: 起始行
        :return: None
        """
        ws = self._wb['analysis']
        bar = BarChart()
        bar.type = 'bar'
        bar.style = 11
        bar.title = '失败接口概况图'
        bar.y_axis.title = '通过或失败用例个数'
        if row_ != 4:
            data = Reference(ws, min_col=2, min_row=3, max_row=row_ - 1, max_col=3)
        else:
            data = Reference(ws, min_col=2, min_row=3, max_row=row_, max_col=3)
        if row_ != 4:
            cats = Reference(ws, min_col=1, min_row=4, max_row=row_ - 1)
        else:
            cats = Reference(ws, min_col=1, min_row=4, max_row=row_)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.shape = 4
        ws.add_chart(bar, "F12")
        bar.width = 30
        bar.height = 0.5 * (row_ + 20)  # 根据行数计算自适应条形图高度
        self._log.info('已生成失败接口概况条形图.')

    def _get_case_num(self):
        """
        计算最大用例数
        :return: None
        """
        ws = self._wb['results']
        case_num = ws.max_row - 1
        self._case_num = case_num
        self._log.info('计算出本次测试用例数：{}'.format(case_num))

    def _get_case_failure_num(self):
        """
        计算失败用例数
        :return: None
        """
        ws = self._wb['results']
        case_failure = 0
        for i in range(2, ws.max_row + 1):
            if ws.cell(i, JUDGEMENT_RESULT_COL).value == 'FAILURE':
                case_failure += 1
        self._case_failure = case_failure
        self._log.info('计算出本次测试失败用例数：{}'.format(case_failure))

    def _get_api_num(self):
        """
        计算接口总数
        :return: None
        """
        ws = self._wb['results']
        api_num = 0
        for i in range(2, ws.max_row + 1):
            if ws.cell(i, CASE_NUMBER_COL).value == 1:
                api_num += 1
        self._api_num = api_num
        self._log.info('计算出本次测试接口数：{}'.format(api_num))

    def _get_api_failure_num(self):
        """
        计算失败接口数
        :return: None
        """
        ws = self._wb['results']
        api_failure = []
        for i in range(2, ws.max_row + 1):
            if ws.cell(i, JUDGEMENT_RESULT_COL).value == 'FAILURE':
                api_failure.append(ws.cell(i, API_NAME_COL).value)
        api_failure = list(set(api_failure))
        self._api_failure = len(api_failure)
        self._log.info('计算出本次测试失败接口数：{}'.format(len(api_failure)))

    def _set_known_cells_value(self):
        """
        先设置可确定位置单元格的值，并处理格式
        :return: None
        """
        ws = self._wb['analysis']

        value_a1 = '分析结果：\n' \
                   '1.本次接口计 {0:^7} 个，测试用例 {1:^7} 个。\n' \
                   '2.失败接口计 {2:^7} 个，通过接口 {3:^7} 个。\n' \
                   '3.失败用例计 {4:^7} 个，通过用例 {5:^7} 个。\n' \
                   '4.接口通过率 {6:^9.2%} ，用例通过率 {7:^9.2%} 。'.format(
                    self._api_num,
                    self._case_num,
                    self._api_failure,
                    self._api_num - self._api_failure,
                    self._case_failure,
                    self._case_num - self._case_failure,
                    (self._api_num - self._api_failure) / self._api_num,
                    (self._case_num - self._case_failure) / self._case_num)

        ws.cell(1, 1).value = value_a1
        ws.cell(2, 1).value = '失败接口概览'
        ws.cell(3, 1).value = '接口名称'
        ws.cell(3, 2).value = '通过'
        ws.cell(3, 3).value = '失败'
        ws.cell(3, 4).value = '总计'
        # 合并单元格，设置字体样式，行高，列宽
        ws.column_dimensions['A'].width = 36
        ws.column_dimensions['B'].width = 19
        ws.column_dimensions['C'].width = 19
        ws.column_dimensions['D'].width = 19
        ws.row_dimensions[1].height = 135
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 20

        ws.merge_cells('A1:D1')
        ws.merge_cells('A2:D2')

        font1 = Font(size=16, bold=True)
        font2 = Font(size=14, bold=True)
        font3 = Font(size=12, bold=True)
        ws.cell(1, 1).font = font1
        ws.cell(2, 1).font = font2
        for i in range(4):
            ws.cell(3, i + 1).font = font3

        fill1 = PatternFill(fill_type='solid', fgColor="00BFFF")
        fill2 = PatternFill(fill_type='solid', fgColor="87CEFA")
        ws.cell(1, 1).fill = fill1
        for i in range(4):
            ws.cell(3, i + 1).fill = fill2

        alignment1 = Alignment(vertical='center')
        alignment2 = Alignment(vertical='center', horizontal='center')
        ws.cell(1, 1).alignment = alignment1
        ws.cell(2, 1).alignment = alignment2

    def _set_unknown_cells_value(self, row_):
        """
        再设置未确定位置单元格的值
        :param row_: 起始行
        :return: None
        """
        ws = self._wb['analysis']

        ws.cell(row_, 1).value = '总计'
        if row_ != 4:
            ws.cell(row_, 2).value = '=SUM(B4:{})'.format('B' + str(row_ - 1))
            ws.cell(row_, 3).value = '=SUM(C4:{})'.format('C' + str(row_ - 1))
            ws.cell(row_, 4).value = '=SUM(D4:{})'.format('D' + str(row_ - 1))
        else:
            ws.cell(row_, 2).value = 0
            ws.cell(row_, 3).value = 0
            ws.cell(row_, 4).value = 0

        for i in range(4):
            ws.cell(row_, i + 1).font = Font(bold=True)

        ws.cell(row_ + 1, 1).value = '失败用例概览'
        ws.cell(row_ + 2, 1).value = '接口名称'
        ws.cell(row_ + 2, 2).value = '用例名称'
        ws.cell(row_ + 2, 3).value = '用例编号'
        ws.cell(row_ + 2, 4).value = '超链接'

        ws.merge_cells('{}:{}'.format(('A' + str(row_ + 1)), ('D' + str(row_ + 1))))

        ws.row_dimensions[row_ + 1].height = 25
        ws.row_dimensions[row_ + 2].height = 20

        font1 = Font(size=14, bold=True)
        font2 = Font(size=12, bold=True)
        ws.cell(row_ + 1, 1).font = font1
        for i in range(4):
            ws.cell(row_ + 2, i + 1).font = font2

        fill = PatternFill(fill_type='solid', fgColor="87CEFA")
        for i in range(4):
            ws.cell(row_ + 2, i + 1).fill = fill

        alignment = Alignment(vertical='center', horizontal='center')
        ws.cell(row_ + 1, 1).alignment = alignment


if __name__ == '__main__':

    # 调用实例
    t = BackupOrNewFile()
    t.backup_result_file()
